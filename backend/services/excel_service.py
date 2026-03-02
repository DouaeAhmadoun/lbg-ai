"""
Excel Shipment File Generator Service
Processes client data and generates market-specific shipment files
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import io
import re
import unicodedata
from typing import Dict, List, Optional, Tuple
import zipfile
from datetime import datetime
from pathlib import Path

# Absolute path to templates directory (robust regardless of working directory)
_DEFAULT_TEMPLATES_DIR = Path(__file__).resolve().parent.parent / "templates"

# Column mapping configuration
# Maps template columns → client data columns (with fallback)
# Format: 'TEMPLATE_COL': ['primary_column', 'fallback_column']
COLUMN_MAPPINGS = {
    'IT': {
        'MEMBER_ID': ['user_id'],
        'NOME': ['nombre', 'name'],           # Essayer 'nombre' d'abord, sinon 'name'
        'COGNOME': ['apellido', 'name'],      # Essayer 'apellido' d'abord, sinon 'name'
        'INDIRIZZO': ['direccion'],           # Juste 'direccion'
        'DETTAGLI': ['complemento_direccion'],
        'CAP': ['codigo_postal', 'postal_code'],  # Essayer espagnol puis anglais
        'CITTÀ ': ['ciudad', 'city'],         # Note: espace dans le header du template
        'PROVINCIA': ['provincia'],
        'TELEFONO': ['telefono'],
        'EMAIL': ['email']
    },
    'FR': {
        'MEMBER_ID': ['user_id'],
        'PRENOM': ['nombre', 'name'],
        'NOM': ['apellido', 'name'],
        'ADRESSE': ['direccion'],
        'COMPLEMENT ADRESSE': ['complemento_direccion'],
        'CP': ['codigo_postal', 'postal_code'],
        'VILLE': ['ciudad', 'city'],
        'REGION': ['provincia'],
        'EMAIL ': ['email'],                  # Note: espace dans le header
        'TÉLEFONO': ['telefono']
    },
    'ES': {
        'MEMBER ID ': ['user_id'],            # Note: espace dans le header
        'NOMBRE': ['nombre', 'name'],
        'APELLIDOS': ['apellido', 'name'],
        'DIRECCIÓN': ['direccion'],
        'DETALLES': ['complemento_direccion'],
        'CODIGO POSTAL': ['codigo_postal', 'postal_code'],
        'CIUDAD': ['ciudad', 'city'],
        'PROVINCIA': ['provincia'],
        'TÉLEFONO': ['telefono'],
        'EMAIL ': ['email']                   # Note: espace dans le header
    }
}

# Template sheet names
TEMPLATE_SHEETS = {
    'IT': 'Template IT',
    'FR': 'Sheet1',
    'ES': 'Sheet1'
}

# Market names for file naming
MARKET_NAMES = {
    'IT': 'Italy',
    'FR': 'France',
    'ES': 'Spain'
}

# Postal code ranges for automatic filtering
POSTAL_CODE_RANGES = {
    'IT': {
        'starts_with': ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9'],
        'length': 5,
        'pattern': r'^\d{5}$'
    },
    'ES': {
        'starts_with': ['0', '1', '2', '3', '4', '5'],
        'length': 5,
        'pattern': r'^[0-5]\d{4}$'
    },
    'FR': {
        'starts_with': ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9'],
        'length': 5,
        'pattern': r'^\d{5}$'
    }
}


# French department codes (first 2-3 digits of postal code)
FRENCH_DEPTS = set(range(1, 96)) | {971, 972, 973, 974, 976}

# Italian province codes (2-letter, uppercase) — unambiguous signal for IT classification
ITALIAN_PROVINCE_CODES = {
    'AG', 'AL', 'AN', 'AO', 'AQ', 'AR', 'AP', 'AT', 'AV',
    'BA', 'BT', 'BL', 'BN', 'BG', 'BI', 'BO', 'BZ', 'BS', 'BR',
    'CA', 'CL', 'CB', 'CE', 'CT', 'CZ', 'CH', 'CO', 'CS', 'CR', 'KR', 'CN',
    'EN', 'FM', 'FE', 'FI', 'FG', 'FC', 'FR',
    'GE', 'GO', 'GR', 'IM', 'IS', 'SP', 'LT', 'LE', 'LC', 'LI', 'LO', 'LU',
    'MC', 'MN', 'MS', 'MT', 'ME', 'MI', 'MO', 'MB',
    'NA', 'NO', 'NU', 'OG', 'OT', 'OR',
    'PD', 'PA', 'PR', 'PV', 'PG', 'PU', 'PE', 'PC', 'PI', 'PT', 'PN', 'PZ', 'PO',
    'RG', 'RA', 'RC', 'RE', 'RI', 'RN', 'RM', 'RO',
    'SA', 'SS', 'SV', 'SI', 'SR', 'SO',
    'TA', 'TE', 'TR', 'TO', 'TP', 'TN', 'TV', 'TS',
    'UD', 'VA', 'VE', 'VB', 'VC', 'VR', 'VV', 'VI', 'VT',
}

# Major Italian cities (lowercase, no accents) — fallback when provincia is missing
ITALIAN_CITIES = {
    'roma', 'milano', 'napoli', 'torino', 'palermo', 'genova', 'bologna',
    'firenze', 'bari', 'catania', 'venezia', 'verona', 'messina', 'padova',
    'trieste', 'taranto', 'brescia', 'prato', 'reggio calabria', 'modena',
    'parma', 'livorno', 'cagliari', 'reggio emilia', 'perugia', 'ravenna',
    'ancona', 'ferrara', 'salerno', 'bergamo', 'trento', 'novara', 'vicenza',
    'lecce', 'pesaro', 'arezzo', 'pescara', 'udine', 'foggia', 'siracusa',
    'sassari', 'monza', 'rimini', 'cosenza', 'piacenza', 'catanzaro',
    'la spezia', 'bolzano', 'terni', 'forlì', 'forli', 'potenza', 'matera',
    'asti', 'alessandria', 'mantova', 'cremona', 'como', 'varese', 'lodi',
    'lecco', 'sondrio', 'brindisi', 'barletta', 'andria', 'benevento',
    'avellino', 'caserta', 'frosinone', 'latina', 'rieti', 'viterbo',
    'grosseto', 'siena', 'pistoia', 'lucca', 'pisa', 'massa', 'carrara',
    'ascoli piceno', 'macerata', 'fermo', 'teramo', 'chieti', 'campobasso',
    'isernia', 'agrigento', 'trapani', 'ragusa', 'nuoro', 'oristano',
}

# Spanish province names (lowercase, no accents) for IT/ES disambiguation
SPANISH_PROVINCE_NAMES = {
    'alava', 'araba', 'albacete', 'alicante', 'alacant', 'almeria', 'avila',
    'badajoz', 'baleares', 'balears', 'illes balears', 'barcelona', 'burgos',
    'caceres', 'cadiz', 'castellon', 'castelló', 'ciudad real', 'cordoba',
    'coruña', 'coruna', 'a coruña', 'la coruña', 'cuenca', 'girona', 'gerona',
    'granada', 'guadalajara', 'guipuzcoa', 'gipuzkoa', 'huelva', 'huesca',
    'jaen', 'leon', 'lleida', 'lerida', 'la rioja', 'rioja', 'lugo', 'madrid',
    'malaga', 'murcia', 'navarra', 'navarre', 'ourense', 'orense', 'palencia',
    'las palmas', 'pontevedra', 'salamanca', 'santa cruz de tenerife', 'tenerife',
    'cantabria', 'segovia', 'sevilla', 'seville', 'soria', 'tarragona', 'teruel',
    'toledo', 'valencia', 'valladolid', 'vizcaya', 'bizkaia', 'zamora',
    'zaragoza', 'asturias', 'ceuta', 'melilla',
}


def _normalize(text: str) -> str:
    """Lowercase and strip accents from a string"""
    return ''.join(
        c for c in unicodedata.normalize('NFD', text.lower().strip())
        if unicodedata.category(c) != 'Mn'
    )


def _classify_row_to_market(row) -> Optional[str]:
    """
    Classify a data row to IT, FR, or ES.

    Priority: province field first (most reliable), then postal code as fallback.
    Rationale: Italian postal codes (00xxx–98xxx) have first-2-digits that overlap
    with French department codes (01–95), making postal-code-only classification
    unreliable. Italian 2-letter province codes (MI, NA, RM, ...) are unambiguous.
    """
    # 1. Province-based classification (highest priority)
    if 'provincia' in row.index and pd.notna(row['provincia']):
        prov_raw = str(row['provincia']).strip()
        if prov_raw and prov_raw not in ('nan', 'None', ''):
            # Italian: exactly 2 uppercase letters matching a known province code
            if re.match(r'^[A-Z]{2}$', prov_raw) and prov_raw in ITALIAN_PROVINCE_CODES:
                return 'IT'
            # Spanish: full province name
            if _normalize(prov_raw) in SPANISH_PROVINCE_NAMES:
                return 'ES'

    # 2. City-based fallback for IT (helps records with missing provincia)
    for col in ('ciudad', 'city'):
        if col in row.index and pd.notna(row[col]):
            city_norm = _normalize(str(row[col]))
            if city_norm in ITALIAN_CITIES:
                return 'IT'
            break

    # 3. Postal code fallback
    postal_code = None
    for col in ('codigo_postal', 'postal_code'):
        if col in row.index and pd.notna(row[col]):
            val = row[col]
            postal_code = str(int(val)).zfill(5) if isinstance(val, (int, float)) else str(val).strip()
            break

    if not postal_code or not re.match(r'^\d{5}$', postal_code):
        return None

    pc_int = int(postal_code)
    dept = pc_int // 1000

    if dept in FRENCH_DEPTS:
        return 'FR'
    if pc_int >= 60000:
        return 'IT'

    return 'IT'  # Default for ambiguous 00000-59999 without province signal


class ShipmentProcessor:
    """Handles shipment file generation"""
    
    def __init__(self, templates_dir: str = None):
        self.client_data = None
        self.templates = {}
        self.templates_dir = templates_dir or str(_DEFAULT_TEMPLATES_DIR)

        # Auto-load latest templates if directory exists
        self._load_latest_templates()
    
    def _load_latest_templates(self):
        """Load the latest template for each market from templates directory"""
        templates_path = Path(self.templates_dir)
        if not templates_path.exists():
            print(f"⚠️  Templates directory not found: {self.templates_dir}")
            templates_path.mkdir(parents=True, exist_ok=True)
            print(f"✅ Created templates directory: {self.templates_dir}")
            return
        
        # Pattern: Shipment_{MARKET}_{YYYYMMDD_HHMMSS}.xlsx
        pattern = re.compile(r'^Shipment_([A-Z]{2})_(\d{8}_\d{6})\.xlsx$')
        
        # Group templates by market
        market_templates = {}
        
        for file_path in templates_path.glob("Shipment_*.xlsx"):
            match = pattern.match(file_path.name)
            if match:
                market = match.group(1)
                timestamp = match.group(2)
                
                if market not in market_templates:
                    market_templates[market] = []
                
                market_templates[market].append({
                    'path': file_path,
                    'timestamp': timestamp,
                    'filename': file_path.name
                })
        
        # Load the latest template for each market
        for market, templates_list in market_templates.items():
            # Sort by timestamp descending (most recent first)
            templates_list.sort(key=lambda x: x['timestamp'], reverse=True)
            latest = templates_list[0]
            
            try:
                with open(latest['path'], 'rb') as f:
                    self.templates[market] = f.read()
                print(f"✅ Loaded latest template for {market}: {latest['filename']}")
            except Exception as e:
                print(f"❌ Error loading {market} template: {e}")
    
    def get_available_markets(self) -> List[str]:
        """Get list of markets with templates loaded"""
        return sorted(list(self.templates.keys()))
    
    def load_client_data(self, file_bytes: bytes) -> pd.DataFrame:
        """Load client data from Excel bytes"""
        try:
            # Try 'Resultado consulta' first, fall back to the first sheet
            xl = pd.ExcelFile(io.BytesIO(file_bytes))
            sheet_name = 'Resultado consulta' if 'Resultado consulta' in xl.sheet_names else xl.sheet_names[0]
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
            print(f"📄 Loaded sheet: '{sheet_name}' ({len(df)} rows, columns: {list(df.columns)})")

            # Only enforce truly non-substitutable columns
            required_columns = ['user_id', 'email']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

            # Clean data
            df = df.fillna('')

            # Convert numeric columns to clean strings (removes .0 suffix from floats)
            numeric_columns = ['user_id', 'codigo_postal', 'postal_code', 'telefono']
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = df[col].astype(str).str.replace('.0', '', regex=False)

            self.client_data = df
            return df
        
        except Exception as e:
            raise ValueError(f"Error loading client data: {str(e)}")
    
    def load_template(self, file_bytes: bytes, market: str) -> openpyxl.Workbook:
        """Load market template"""
        try:
            workbook = load_workbook(io.BytesIO(file_bytes))
            self.templates[market] = file_bytes
            return workbook
        except Exception as e:
            raise ValueError(f"Error loading {market} template: {str(e)}")
    
    def auto_filter_by_market(self, market: str) -> pd.DataFrame:
        """Filter clients for a specific market using mutual-exclusion classification"""
        if self.client_data is None:
            return pd.DataFrame()

        mask = self.client_data.apply(
            lambda row: _classify_row_to_market(row) == market, axis=1
        )
        filtered = self.client_data[mask]
        print(f"✂️  {market}: {len(filtered)} records after filtering (from {len(self.client_data)})")
        return filtered
    
    def manual_filter_by_ids(self, user_ids: List[int]) -> pd.DataFrame:
        """Manually filter clients by user IDs"""
        if self.client_data is None:
            return pd.DataFrame()
        
        return self.client_data[
            self.client_data['user_id'].astype(str).isin([str(uid) for uid in user_ids])
        ]
    
    def generate_shipment_file(
        self,
        market: str,
        filtered_data: pd.DataFrame
    ) -> bytes:
        """Generate shipment file for a market"""
        if market not in self.templates:
            raise ValueError(f"Template for {market} not loaded")

        workbook = load_workbook(io.BytesIO(self.templates[market]))
        sheet_name = TEMPLATE_SHEETS[market]
        worksheet = workbook[sheet_name]
        mapping = COLUMN_MAPPINGS[market]

        # Build header → column index from row 1, stripping spaces for robust matching
        header_to_col = {
            str(cell.value).strip(): cell.column
            for cell in worksheet[1]
            if cell.value is not None
        }

        # Clear any pre-existing data rows from the template (keep header row 1)
        if worksheet.max_row > 1:
            for row_cells in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row_cells:
                    cell.value = None

        print(f"📝 Writing {len(filtered_data)} rows to {market} template")

        for seq_idx, (_, row) in enumerate(filtered_data.iterrows()):
            excel_row = 2 + seq_idx  # row 1 is header

            for template_col, source_cols in mapping.items():
                col_idx = header_to_col.get(template_col.strip())
                if col_idx is None:
                    continue

                value = ''
                for source_col in source_cols:
                    if source_col in row:
                        potential_value = row[source_col]
                        if potential_value is None or pd.isna(potential_value):
                            continue
                        str_value = str(potential_value).strip()
                        if str_value in ('', 'None', 'nan', 'NaN'):
                            continue
                        value = str(potential_value).replace('.0', '') if isinstance(potential_value, (int, float)) else str_value
                        break

                worksheet.cell(row=excel_row, column=col_idx, value=value)

        print(f"✅ {market}: Wrote {len(filtered_data)} rows successfully")

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
    
    def generate_all_files(
        self,
        markets: List[str],
        filter_mode: str = "auto",
        manual_filters: Dict[str, List[int]] = None
    ) -> Dict[str, bytes]:
        """Generate files for multiple markets"""
        generated_files = {}
        
        print(f"🔍 generate_all_files called for markets: {markets}")
        print(f"📊 Total client data records: {len(self.client_data) if self.client_data is not None else 0}")
        
        if self.client_data is None or self.client_data.empty:
            print(f"❌ No client data available")
            return generated_files
        
        for market in markets:
            try:
                if filter_mode == "manual" and manual_filters and market in manual_filters:
                    filtered_data = self.manual_filter_by_ids(manual_filters[market])
                else:
                    filtered_data = self.auto_filter_by_market(market)

                print(f"📋 {market}: Generating with {len(filtered_data)} filtered records")

                file_bytes = self.generate_shipment_file(market, filtered_data)
                generated_files[market] = file_bytes
                print(f"✅ {market}: File generated successfully")
            
            except Exception as e:
                print(f"❌ {market}: Error - {str(e)}")
                import traceback
                traceback.print_exc()
                raise ValueError(f"Error generating {market} file: {str(e)}")
        
        print(f"📦 Final generated_files: {list(generated_files.keys())}")
        return generated_files


def create_zip_file(files_dict: Dict[str, bytes], timestamp: str = None) -> bytes:
    """Create ZIP file containing all shipment files"""
    from datetime import datetime
    
    if timestamp is None:
        timestamp = datetime.now().strftime('%Y-%m-%d_%Hh%M')
    
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for market, file_bytes in files_dict.items():
            market_name = MARKET_NAMES.get(market, market)  # Italy, France, Spain
            filename = f"Shipment_{market_name}_{timestamp}.xlsx"
            zip_file.writestr(filename, file_bytes)
    
    zip_buffer.seek(0)
    return zip_buffer.getvalue()


def detect_markets(data: pd.DataFrame) -> Dict[str, int]:
    """
    Automatically detect markets based on postal codes and province names.
    Returns: {'IT': 15, 'FR': 8, 'ES': 12} - number of records per market
    """
    market_counts = {'IT': 0, 'FR': 0, 'ES': 0}
    for _, row in data.iterrows():
        market = _classify_row_to_market(row)
        if market:
            market_counts[market] += 1
    return market_counts


def validate_shipment_data(data: pd.DataFrame, market: str) -> Dict:
    """
    Validate shipment data and return a detailed report
    Returns: {
        'total_rows': int,
        'valid_rows': int,
        'blocking_errors': [{row, user_id, issue}, ...],
        'warnings': [{row, user_id, issue}, ...],
        'validations': [{row, user_id, issue}, ...],
        'summary': str
    }
    """
    import re
    
    blocking_errors = []
    warnings = []
    validations = []
    
    # Postal code patterns by market
    postal_patterns = {
        'IT': r'^\d{5}$',
        'FR': r'^\d{5}$',
        'ES': r'^[0-5]\d{4}$'
    }
    
    for idx, row in data.iterrows():
        row_num = idx + 2  # Excel row (header is 1, data starts at 2)
        user_id = row.get('user_id', 'Unknown')
        
        # Helper function to get value with fallback
        def get_value(primary_col, fallback_col=None):
            val = row.get(primary_col)
            if pd.notna(val) and str(val).strip() not in ['', 'None', 'nan']:
                return str(val).strip()
            if fallback_col:
                val = row.get(fallback_col)
                if pd.notna(val) and str(val).strip() not in ['', 'None', 'nan']:
                    return str(val).strip()
            return None
        
        # 1. BLOCKING ERRORS - Prevent delivery
        
        # Missing address
        address = get_value('direccion')
        if not address:
            blocking_errors.append({
                'row': row_num,
                'user_id': user_id,
                'issue': 'Missing address'
            })
        
        # Missing postal code
        postal_code = get_value('codigo_postal', 'postal_code')
        if not postal_code:
            blocking_errors.append({
                'row': row_num,
                'user_id': user_id,
                'issue': 'Missing postal code'
            })
        
        # Missing city
        city = get_value('ciudad', 'city')
        if not city:
            blocking_errors.append({
                'row': row_num,
                'user_id': user_id,
                'issue': 'Missing city'
            })
        
        # Missing name
        name = get_value('nombre', 'name')
        if not name:
            blocking_errors.append({
                'row': row_num,
                'user_id': user_id,
                'issue': 'Missing name'
            })
        
        # Missing email
        email = get_value('email')
        if not email:
            blocking_errors.append({
                'row': row_num,
                'user_id': user_id,
                'issue': 'Missing email'
            })
        
        # 2. WARNINGS - Delivery possible but risky
        
        # Missing phone
        phone = get_value('telefono')
        if not phone:
            warnings.append({
                'row': row_num,
                'user_id': user_id,
                'issue': 'Missing phone number'
            })
        
        # Missing province/region
        province = get_value('provincia')
        if not province:
            warnings.append({
                'row': row_num,
                'user_id': user_id,
                'issue': 'Missing province/region'
            })
        
        # Missing last name
        lastname = get_value('apellido')
        if not lastname and name:
            warnings.append({
                'row': row_num,
                'user_id': user_id,
                'issue': 'Missing last name (only first name provided)'
            })
        
        # 3. VALIDATIONS - Suspicious data
        
        # Invalid postal code format (only check if not a valid number)
        if postal_code:
            pattern = postal_patterns.get(market)
            # Don't validate if it's a number with .0 (will be cleaned)
            if pattern and not re.match(pattern, postal_code):
                # Check if it's NOT a float-like string
                try:
                    float_val = float(postal_code)
                    # If it's a valid number, skip validation (will be cleaned)
                except:
                    # Not a number, validate format
                    validations.append({
                        'row': row_num,
                        'user_id': user_id,
                        'issue': f'Invalid postal code format: "{postal_code}"'
                    })
        
        # Invalid email format
        if email and '@' not in email:
            validations.append({
                'row': row_num,
                'user_id': user_id,
                'issue': f'Invalid email format: "{email}"'
            })
        
        # Phone too short
        if phone and len(re.sub(r'\D', '', phone)) < 7:
            validations.append({
                'row': row_num,
                'user_id': user_id,
                'issue': f'Phone number too short: "{phone}"'
            })
        
        # Address too short
        if address and len(address) < 5:
            validations.append({
                'row': row_num,
                'user_id': user_id,
                'issue': f'Address too short: "{address}"'
            })
    
    total_rows = len(data)
    valid_rows = total_rows - len(set(e['row'] for e in blocking_errors))
    
    return {
        'total_rows': total_rows,
        'valid_rows': valid_rows,
        'blocking_errors': blocking_errors,
        'warnings': warnings,
        'validations': validations,
        'has_issues': len(blocking_errors) > 0 or len(warnings) > 0 or len(validations) > 0
    }


def get_market_preview(
    processor: ShipmentProcessor,
    market: str,
    filter_mode: str = "auto"
) -> Dict:
    """Get preview of data that would be exported for a market"""
    if filter_mode == "auto":
        filtered_data = processor.auto_filter_by_market(market)
    else:
        filtered_data = processor.client_data
    
    return {
        "market": market,
        "total_records": len(filtered_data),
        "preview": filtered_data.head(10).to_dict('records') if not filtered_data.empty else []
    }